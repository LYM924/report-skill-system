"""FAQ 路由：列表/详情/保存/删除/导入/归属建议/相似推荐/浏览计数

数据流约定（修复后）：
  - 保存：写一份 .md 文件 + 写 faqs 表（save_faq write_file=False 单文件） + 关键词双表 + 重建内存索引
  - 删除：物理删文件（限 data/faq/ 内）+ faqs 表软删除 + 重建内存索引
  - 向量索引在变更后由后台线程全量重建（避免 BM25/向量脱节）
"""
import datetime
import json
import logging
import re

from fastapi import APIRouter, BackgroundTasks, Depends, Query, UploadFile
from fastapi.responses import JSONResponse
from pydantic import BaseModel

import main
from auth import verify_token
from config import settings
from repository import get_repo
from repository.base import FAQ
from repository.dept_mapping import get_dept_path, get_submodule_path
from service.paths import safe_data_path
from routes.health import record_write_failure

router = APIRouter(tags=["FAQ"])

DEPT_CODES = {"数智财务组": "SZ", "免疫规划组": "YM", "电子档案组": "DZ", "数字化支撑组": "ZH"}
STATUS_MAP = {"active": 1, "outdated": 2, "deprecated": 3, "draft": 0}


class ImportBody(BaseModel):
    pass  # 导入走 multipart，占位


def _reload_after_faq_change(rebuild_vector: bool = True):
    """FAQ 变更后：重建 FAQ 内存结构 + BM25 + 可选后台重建向量索引 + 失效缓存"""
    if main.search_engine is not None:
        main.search_engine.after_faq_change(rebuild_vector=rebuild_vector)
    # 失效相关缓存（保证写后立即读一致）
    from service.cache import cache_delete, FAQ_WRITE_KEYS
    cache_delete(*FAQ_WRITE_KEYS)


def _resolve_kw_ids(dept: str, sub_module: str, repo,
                    dept_id: int = 0, module_id: int = 0):
    """解析关键词写入所需的 module_id/dept_id（外键，NULL 安全）

    前端携带的 ID 优先（唯一标识）；未携带时按名称解析——dept_id 按用户
    选择的部门名（不取模块行——modules 表部门关联是旧组织架构）；
    module_id 优先「部门+名称」联合匹配，无匹配回退名称唯一查找。
    """
    resolved_dept = dept_id or repo.resolve_dept_id(dept)
    resolved_module = module_id or (
        repo.resolve_module_id(sub_module, dept_name=dept) if sub_module else None
    )
    return resolved_module, resolved_dept


@router.get("/faq")
async def list_faqs(id: str = Query(""), user: str = Depends(verify_token)):
    """FAQ 列表 / 详情（DB 直查，保证与数据库实时一致）

    列表不再遍历内存 faq_docs，直接查 faqs 表；
    详情优先读文件，文件缺失回退读 DB faq_answer。
    """
    repo = get_repo()

    if id:
        # 详情：先从 DB 查元数据
        row = repo._execute_one(
            "SELECT faq_code, faq_title, dept, sub_module, module, tags, file_path, faq_answer "
            "FROM faqs WHERE faq_code = :code AND is_deleted = FALSE",
            {"code": id}
        )
        if not row:
            return JSONResponse({"error": "FAQ 不存在"}, status_code=404)

        content = ""
        file_path = row.get("file_path", "")
        p = safe_data_path(file_path) if file_path else None
        if p and p.exists():
            raw = p.read_text(encoding="utf-8")
            if raw.startswith("---"):
                parts = raw.split("---", 2)
                if len(parts) >= 3:
                    content = parts[2].lstrip("\n")
                else:
                    content = raw
            else:
                content = raw
        else:
            # 文件缺失时回退读 DB 的 faq_answer
            if row.get("faq_answer"):
                content = row["faq_answer"]
            else:
                logging.getLogger("faq").warning("FAQ 文件和 DB 均缺失正文: %s", id)

        # tags 可能是 list 或逗号分隔字符串
        tags = row.get("tags") or []
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(",") if t.strip()]

        return {
            "title": row.get("faq_title", ""),
            "dept": row.get("dept", ""),
            "sub_module": row.get("sub_module", ""),
            "module": row.get("module", ""),
            "keywords": tags,
            "path": file_path,
            "content": content,
            "id": id,
        }

    # 列表：DB 直查
    rows = repo.get_faqs_page()
    faqs = []
    for row in rows:
        tags = row.get("tags") or []
        if isinstance(tags, str):
            tags = [t.strip() for t in tags.split(",") if t.strip()]
        faqs.append({
            "id": row.get("faq_code", ""),
            "title": row.get("faq_title", ""),
            "dept": row.get("dept", ""),
            "sub_module": row.get("sub_module", ""),
            "module": row.get("module", ""),
            "keywords": tags,
            "path": row.get("file_path", ""),
        })
    return {"faqs": faqs}


@router.get("/faq/delete")
async def delete_faq(path: str = Query(...), user: str = Depends(verify_token)):
    """删除 FAQ：DB 软删除（主）+ 物理删文件（辅，文件已不存在也继续）+ 重建索引

    修复：不再先检查文件是否存在再决定能否删除。
    数据源以 DB 为主——DB 有记录就能软删除，文件丢失只是日志告警不阻断。
    """
    p = safe_data_path(path)
    if not p or not str(p).startswith(str(settings.DATA_DIR / "faq")):
        return JSONResponse({"error": "路径非法"}, status_code=400)

    # DB 软删除（主操作）
    repo = get_repo()
    try:
        result = repo.delete_faq(path)
        if not result.get("ok"):
            return JSONResponse({"error": result.get("error", "FAQ 不存在")}, status_code=404)
    except Exception:
        record_write_failure("faq_delete")
        return JSONResponse({"error": "FAQ 删除失败（数据库）"}, status_code=500)

    # 物理删文件（辅操作，文件已不存在也正常）
    if p.exists():
        p.unlink()
    else:
        import logging
        logging.getLogger("faq").warning(f"FAQ 文件已缺失，仅 DB 软删除: {path}")

    _reload_after_faq_change()
    return {"ok": True, "message": "已删除"}


@router.get("/faq/suggest")
async def faq_suggest(title: str = Query(""), keywords: str = Query(""),
                      user: str = Depends(verify_token)):
    """FAQ 归属建议：对关键词索引投票得出最可能的部门/模块"""
    if main.search_engine is None:
        return {"dept": "", "module": "", "dept_votes": [], "module_votes": []}

    tokens = [t.strip() for t in (title + " " + keywords).split(",") if t.strip()]
    if title and not keywords:
        tokens = [title] + [t.strip() for t in title.split() if len(t.strip()) >= 2]
    dept_votes, module_votes = {}, {}
    for token in tokens:
        if not token:
            continue
        for kw, entries in main.search_engine.keyword_map.items():
            if kw in token or token in kw:
                for e in entries[:3]:
                    d, m = e.get("dept", ""), e.get("module", "")
                    if d:
                        dept_votes[d] = dept_votes.get(d, 0) + 1
                    if m:
                        module_votes[m] = module_votes.get(m, 0) + 1

    best_dept = max(dept_votes, key=dept_votes.get) if dept_votes else "数智财务组"
    best_module = max(module_votes, key=module_votes.get) if module_votes else "浙里报"
    return {
        "dept": best_dept,
        "module": best_module,
        "dept_votes": sorted(dept_votes.items(), key=lambda x: -x[1])[:5],
        "module_votes": sorted(module_votes.items(), key=lambda x: -x[1])[:5],
    }


@router.get("/faq/similar")
async def faq_similar(keywords: str = Query(""), user: str = Depends(verify_token)):
    """相似 FAQ 推荐：标题命中+3 / 关键词命中+2，取前 5"""
    if main.search_engine is None:
        return {"faqs": []}
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    if not kw_list:
        return {"faqs": []}
    scored = []
    for faq in main.search_engine.faq_docs:
        score = 0
        for kw in kw_list:
            if kw in faq.get("title", ""):
                score += 3
            for fkw in faq.get("keywords", []):
                if kw in fkw or fkw in kw:
                    score += 2
        if score > 0:
            scored.append({
                "id": faq.get("faq_id", ""),
                "title": faq.get("title", ""),
                "keywords": faq.get("keywords", []),
                "dept": faq.get("dept", ""),
                "score": score,
            })
    scored.sort(key=lambda x: -x["score"])
    return {"faqs": scored[:5]}


@router.get("/faq/save")
async def save_faq(
    id: str = Query(""),
    title: str = Query(""),
    keywords: str = Query(""),
    dept: str = Query(""),
    sub_module: str = Query(""),
    module: str = Query(""),
    content: str = Query(""),
    status: str = Query("draft"),
    dept_id: int = Query(0),
    module_id: int = Query(0),
    user: str = Depends(verify_token),
):
    """保存 FAQ：单文件 + faqs 表 + 关键词双表 + 重建索引

    带 id = 更新已有 FAQ（按 faq_code 幂等 upsert）；不带 id = 新增并生成 FAQ-{部门}-{模块}-{NNN}
    dept_id/module_id 优先于名称使用（前端选择器携带唯一 ID），传 ID 未传名称时反查名称。
    """
    repo = get_repo()
    if dept_id and not dept:
        dept = repo.get_department_name(dept_id)
    if module_id and not (sub_module or module):
        sub_module = repo.get_module_name(module_id)
        module = sub_module or module
    if not title or not dept:
        return JSONResponse({"error": "title 和 dept 为必填参数"}, status_code=422)

    from keyword_extractor import get_extractor, build_extractor_idf

    dept_path = get_dept_path(dept) or "other"
    sub_path = get_submodule_path(sub_module) if sub_module else ""
    faq_dir = settings.DATA_DIR / "faq" / dept_path / sub_path
    faq_dir.mkdir(parents=True, exist_ok=True)

    # 解析最终 dept_id/module_id（前端 ID 优先，缺失时按名称回退）
    resolved_d_id = dept_id or repo.resolve_dept_id(dept)
    resolved_m_id = module_id or repo.resolve_module_id(sub_module or module, dept_name=dept)

    # 生成/复用 FAQ 编码
    if not id:
        dept_code = repo.get_dept_code(dept) or DEPT_CODES.get(dept, "XX")
        mod_code = sub_module[:3] if sub_module else "XXX"
        existing = list(faq_dir.glob("*.md"))
        id = f"FAQ-{dept_code}-{mod_code}-{len(existing) + 1:03d}"

    # 关键词：显式提供 → 拆分；否则自动提取
    kw_list = [k.strip() for k in keywords.split(",") if k.strip()]
    if not kw_list and content:
        try:
            extractor = get_extractor()
            if not extractor._built:
                build_extractor_idf(str(settings.DATA_DIR))
            kw_list = extractor.extract(content, top_k=10)
        except Exception:
            kw_list = []

    # 保留既有 related/tickets（编辑场景）
    related, tickets = [], []
    old_file = faq_dir / f"{id}.md"
    if old_file.exists():
        try:
            old_raw = old_file.read_text(encoding="utf-8")
            for line in old_raw.split("\n"):
                if line.startswith("related:"):
                    try:
                        related = json.loads(line.split(":", 1)[1].strip() or "[]")
                    except Exception:
                        pass
                if line.startswith("tickets:"):
                    try:
                        tickets = json.loads(line.split(":", 1)[1].strip() or "[]")
                    except Exception:
                        pass
        except Exception:
            pass

    # 内容清洗：剥离夹带的 frontmatter
    safe_content = content
    if safe_content.strip().startswith("---"):
        parts = safe_content.split("---", 2)
        if len(parts) >= 3:
            safe_content = parts[2].lstrip("\n")

    today = datetime.date.today().isoformat()
    file_content = f"""---
id: {id}
title: {title}
keywords: {json.dumps(kw_list, ensure_ascii=False)}
module: {module}
dept: {dept}
sub_module: {sub_module}
scene: ""
status: {status}
version_from: ""
created: {today}
reviewed: {today}
related: {json.dumps(related, ensure_ascii=False)}
tickets: {json.dumps(tickets, ensure_ascii=False)}
---

# {title}

{safe_content}
"""
    file_path = faq_dir / f"{id}.md"
    file_path.write_text(file_content, encoding="utf-8")
    rel_path = str(file_path.relative_to(settings.PROJECT_DIR))

    # 写 faqs 表（write_file=False：文件已由本路由写入，避免双文件）
    faq = FAQ(
        faq_code=id, faq_title=title, faq_question=title, faq_answer=safe_content,
        content=file_content, path=rel_path, tags=kw_list, dept=dept,
        dept_id=resolved_d_id or 0, sub_module=sub_module, module=module,
        module_id=resolved_m_id or 0, scene="",
        status=STATUS_MAP.get(status, 0), sort_num=0, view_count=0,
        source_file_name=f"{id}.md", version_from="",
        related=related, tickets=tickets,
        update_time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        is_deleted=0,
    )
    try:
        repo.save_faq(faq, write_file=False)
    except Exception as e:
        record_write_failure("faq_save")
        # 文件已写入，DB 失败不阻断（启动 bulk_import 会兜底回导）
        return {"ok": True, "faq_id": id, "path": rel_path,
                "warning": f"数据库写入失败（已写入文件）: {e}"}

    # 关键词双表写入（复活式 upsert）
    try:
        m_id, d_id = _resolve_kw_ids(dept, sub_module, repo,
                                     dept_id=resolved_d_id, module_id=resolved_m_id)
        for kw in kw_list:
            repo.add_keyword(kw, m_id or 0, d_id or 0, dept, kb_path=rel_path)
    except Exception:
        record_write_failure("keyword_write")

    _reload_after_faq_change()
    return {"ok": True, "faq_id": id, "path": rel_path}


@router.post("/faq/import")
@router.get("/faq/import")
async def faq_import(
    file: UploadFile = None,
    background_tasks: BackgroundTasks = None,
    user: str = Depends(verify_token),
):
    """FAQ Excel 批量导入（multipart）：写 .md 文件 + 回导 faqs 表 + 重建索引"""
    if file is None:
        return JSONResponse({"error": "需要上传 Excel 文件（multipart file 字段）"}, status_code=422)
    try:
        import openpyxl
        from io import BytesIO
        data = await file.read()
        wb = openpyxl.load_workbook(BytesIO(data))
        ws = wb.active
    except Exception as e:
        return JSONResponse({"error": f"Excel 解析失败: {e}"}, status_code=400)

    # 列名映射
    header = [str(c.value or "").strip() for c in ws[1]]
    col_map = {}
    for idx, name in enumerate(header):
        if name in ("标题", "问题", "title"):
            col_map.setdefault("title", idx)
        elif name in ("关键词", "keywords", "标签", "tags"):
            col_map.setdefault("keywords", idx)
        elif name in ("部门", "dept", "业务组"):
            col_map.setdefault("dept", idx)
        elif name in ("模块", "module", "子模块"):
            col_map.setdefault("module", idx)
        elif name in ("描述", "问题描述", "description", "现象"):
            col_map.setdefault("desc", idx)
        elif name in ("解决", "方案", "solution", "答案", "answer", "方法"):
            col_map.setdefault("solution", idx)
    if "title" not in col_map:
        return JSONResponse({"error": "缺少标题列"}, status_code=422)

    success, fail = 0, 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        title = str(row[col_map["title"]] or "").strip() if col_map["title"] < len(row) else ""
        if len(title) < 4 or title.isdigit():
            fail += 1
            continue
        dept = str(row[col_map.get("dept", 0)] or "数智财务组").strip() if col_map.get("dept", 0) < len(row) else "数智财务组"
        module = str(row[col_map.get("module", 0)] or "").strip() if col_map.get("module", 0) < len(row) else ""
        kws = str(row[col_map.get("keywords", 0)] or "").strip() if col_map.get("keywords", 0) < len(row) else ""
        desc = str(row[col_map.get("desc", 0)] or "").strip() if col_map.get("desc", 0) < len(row) else ""
        solution = str(row[col_map.get("solution", 0)] or "").strip() if col_map.get("solution", 0) < len(row) else ""

        dept_path = get_dept_path(dept) or "other"
        sub_path = get_submodule_path(module) if module else ""
        faq_dir = settings.DATA_DIR / "faq" / dept_path / sub_path
        faq_dir.mkdir(parents=True, exist_ok=True)
        safe_title = title.replace("/", "-").replace("?", "").replace(":", "")
        kw_list = [k.strip() for k in kws.split(",") if k.strip()]
        existing = list(faq_dir.glob("*.md"))
        faq_code = f"FAQ-{DEPT_CODES.get(dept, 'XX')}-{module[:3] if module else 'XXX'}-{len(existing) + 1:03d}"
        today = datetime.date.today().isoformat()
        content_md = f"""---
id: {faq_code}
title: {title}
keywords: {json.dumps(kw_list, ensure_ascii=False)}
module: {module}
dept: {dept}
sub_module: {module}
scene: ""
status: active
version_from: ""
created: {today}
reviewed: {today}
related: []
tickets: []
---

# {title}

## 问题描述
{desc}

## 原因分析

## 解决方法
{solution}
"""
        (faq_dir / f"{safe_title}.md").write_text(content_md, encoding="utf-8")
        success += 1

    # 回导 faqs 表（幂等）+ 重建索引
    repo = get_repo()
    try:
        repo.bulk_import_faqs()
    except Exception:
        record_write_failure("faq_save")
    _reload_after_faq_change()
    return {"ok": True, "success": success, "fail": fail}


@router.get("/faq/view")
async def faq_view(id: str = Query(...), user: str = Depends(verify_token)):
    """FAQ 浏览次数 +1"""
    repo = get_repo()
    try:
        repo.increment_faq_views(id)
    except Exception:
        record_write_failure("faq_save")
    return {"ok": True}
