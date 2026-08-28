#!/usr/bin/env python3
"""
从 Excel 导入 FAQ 到知识库系统。

读取 【疫苗一体化 FAQ】一线收集.xlsx 中的 FAQ 数据，
导入到 data/faq/immunization/ 目录和 SQLite 数据库。

用法:
  python3 src/server/import_faq_from_excel.py
  python3 src/server/import_faq_from_excel.py --dry-run
"""

import json
import re
import sqlite3
import sys
from pathlib import Path
from datetime import datetime

HERE = Path(__file__).resolve().parent
PROJECT_DIR = HERE.parent.parent  # knowledge-base/
DATA_DIR = PROJECT_DIR / "data"
RUNTIME_DIR = PROJECT_DIR / "runtime"
DB_PATH = RUNTIME_DIR / "knowledge.db"
FAQ_DIR = DATA_DIR / "faq" / "immunization"

EXCEL_FILE = Path.home() / "Downloads" / "【疫苗一体化 FAQ】一线收集.xlsx"

# 模块名 → 英文目录名映射
MODULE_EN_MAP = None  # 从 dept_mapping 动态加载，无需手动维护


def get_en_dir(module_name, db=None):
    """获取模块对应的英文目录名（优先 dept_mapping，其次数据库，最后拼音）"""
    # 1. dept_mapping 映射表
    try:
        from repository.dept_mapping import get_submodule_path
        path = get_submodule_path(module_name)
        if path and path != module_name:
            return path
    except ImportError:
        pass

    # 2. 数据库查询模块的 path 字段
    if db:
        row = db.execute("SELECT path FROM modules WHERE name = ?", (module_name,)).fetchone()
        if row and row["path"]:
            parts = row["path"].split("/")
            if len(parts) >= 2:
                return parts[-1]

    # 3. 拼音兜底
    try:
        from pypinyin import lazy_pinyin
        return "-".join(lazy_pinyin(module_name))
    except ImportError:
        return module_name.lower().replace(" ", "-")


def parse_excel(filepath):
    """解析 Excel，返回 FAQ 列表"""
    try:
        import openpyxl
    except ImportError:
        print("请安装 openpyxl: pip install openpyxl")
        return []

    wb = openpyxl.load_workbook(filepath, data_only=True)
    faqs = []

    # Sheet 1: 主 FAQ 表
    ws = wb["FAQ"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        seq, module, question, answer, image, update_date = (list(row) + [None]*6)[:6]
        if not question or not str(question).strip():
            continue
        question = str(question).strip()
        answer = str(answer).strip() if answer else ""
        module = str(module).strip() if module else "预防接种"

        if len(question) < 5 or len(answer) < 10:
            continue

        faqs.append({
            "module": module,
            "question": question,
            "answer": answer,
            "source": "FAQ sheet",
        })

    # Sheet 2: 免疫程序
    ws = wb["免疫程序（智能机器人）"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        q, a, _ = (list(row) + [None]*3)[:3]
        if not q or not a:
            continue
        q, a = str(q).strip(), str(a).strip()
        if len(q) < 5 or len(a) < 10:
            continue
        faqs.append({
            "module": "预防接种",
            "question": q,
            "answer": a,
            "source": "免疫程序",
        })

    # Sheet 3: 预防接种系统操作FAQ
    ws = wb["预防接种系统操作FAQ"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        cells = list(row) + [None]*7
        q = str(cells[2]).strip() if cells[2] else ""
        a = str(cells[3]).strip() if cells[3] else ""
        if not q or not a or len(q) < 5 or len(a) < 10:
            continue
        faqs.append({
            "module": "预防接种",
            "question": q,
            "answer": a,
            "source": "预防接种系统操作FAQ",
        })

    wb.close()
    return faqs


def import_faqs(faqs, dry_run=False):
    """导入 FAQ 到数据库和文件系统"""
    db = sqlite3.connect(str(DB_PATH))
    db.row_factory = sqlite3.Row

    # 获取 department_id for 免疫规划组
    dept_row = db.execute("SELECT id FROM departments WHERE name = '免疫规划组'").fetchone()
    dept_id = dept_row["id"] if dept_row else 2

    imported = 0
    skipped = 0

    for faq in faqs:
        module_name = faq["module"]
        question = faq["question"]
        answer = faq["answer"]

        # 生成 FAQ ID
        en_mod = get_en_dir(module_name, db)
        safe_title = question[:40].replace("/", "-").replace(":", "-").replace("?", "").replace("？", "")

        # 检查是否已存在
        existing = db.execute(
            "SELECT id FROM faqs WHERE faq_question = ? AND is_deleted = 0",
            (question,)
        ).fetchone()
        if existing:
            skipped += 1
            continue

        # 获取 module_id
        mod_row = db.execute("SELECT id FROM modules WHERE name = ?", (module_name,)).fetchone()
        module_id = mod_row["id"] if mod_row else None

        # 生成 FAQ code
        count = db.execute("SELECT COUNT(*) FROM faqs").fetchone()[0]
        faq_code = f"FAQ-YM-{en_mod.upper()}-{count + 1:03d}"

        # 构建完整内容
        content = f"""# {question}

## 问题描述

{question}

## 解决方法

{answer}

## 排查要点

1. 确认用户描述的问题场景是否与上述一致
2. 按照解决方案步骤逐一排查
3. 如问题未解决，提交技术支持工单

## 关联知识

- 查看 [免疫规划组知识库](../../knowledge/immunization/) 了解相关模块文档
"""

        if not dry_run:
            # 保存到文件
            en_dir = MODULE_EN_MAP.get(module_name, module_name)
            target_dir = FAQ_DIR / en_dir
            target_dir.mkdir(parents=True, exist_ok=True)
            filename = f"{faq_code}.md"
            filepath = target_dir / filename

            md_content = f"""---
id: {faq_code}
title: {safe_title}
keywords: {json.dumps([module_name], ensure_ascii=False)}
module: {module_name}
dept: 免疫规划组
sub_module: {module_name}
scene: 预防接种
status: active
version_from: ""
created: {datetime.now().strftime('%Y-%m-%d')}
reviewed: {datetime.now().strftime('%Y-%m-%d')}
related: []
tickets: []
---

{content}
"""
            filepath.write_text(md_content, encoding="utf-8")

            # 保存到数据库
            rel_path = str(filepath.relative_to(PROJECT_DIR))
            db.execute("""
                INSERT INTO faqs (faq_code, faq_title, faq_question, faq_answer,
                    content, dept, sub_module, module, department_id, module_id,
                    scene, status, tags, file_path, source_file_name, create_time, update_time)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?, ?, ?, ?)
            """, (
                faq_code, safe_title, question, answer, content,
                "免疫规划组", module_name, module_name, dept_id, module_id,
                "预防接种", json.dumps([module_name], ensure_ascii=False),
                rel_path, filename, datetime.now().isoformat(), datetime.now().isoformat()
            ))

        imported += 1

    if not dry_run:
        db.commit()

    print(f"\n{'[DRY RUN] ' if dry_run else ''}导入结果:")
    print(f"  总计: {len(faqs)} 条")
    print(f"  导入: {imported}")
    print(f"  跳过(重复): {skipped}")

    db.close()
    return imported


def main():
    dry_run = "--dry-run" in sys.argv

    if not EXCEL_FILE.exists():
        print(f"❌ 文件不存在: {EXCEL_FILE}")
        return

    print(f"{'[DRY RUN] ' if dry_run else ''}读取: {EXCEL_FILE}")
    faqs = parse_excel(EXCEL_FILE)
    print(f"  解析到 {len(faqs)} 条 FAQ")

    import_faqs(faqs, dry_run)

    if dry_run:
        print("\n💡 使用 python3 src/server/import_faq_from_excel.py 执行实际导入")


if __name__ == "__main__":
    main()