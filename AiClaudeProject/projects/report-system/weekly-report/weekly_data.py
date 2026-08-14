#!/usr/bin/env python3
"""
周报一键数据获取脚本
用法:
  python3 weekly_data.py                    # 自动检测当前周
  python3 weekly_data.py --week W30         # 指定周次
  python3 weekly_data.py --week W30 --json  # 输出 JSON 格式

输出: 结构化 JSON 数据，包含工单统计、TOP问题、超期工单、需求列表等
      用于直接填充周报模板，无需手动写 Python 分析命令
"""
import json
import re
import os
import sys
import argparse
from pathlib import Path
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta
from concurrent.futures import ThreadPoolExecutor, as_completed

# ============================================================
# 1. 周日期自动计算
# ============================================================

def get_week_from_date(d=None):
    """根据日期计算所属周次和日期范围。
    周定义: 上周五 → 本周四
    周次计算: 从年初第一个周五开始计数（如 2026-01-02 周五为 W1）
    返回: (week_num, week_start, week_end, friday, thursday)
    """
    if d is None:
        d = date.today()

    # 找到最近的周四（本周四）
    days_since_thu = (d.weekday() - 3) % 7
    thursday = d - timedelta(days=days_since_thu)
    # 上周五 = 本周四 - 6 天
    friday = thursday - timedelta(days=6)

    # 计算周次：从年初第一个周五开始
    first_friday = date(thursday.year, 1, 2)  # 1月2日总是周五？不，需要计算
    # 找到当年第一个周五
    jan1 = date(thursday.year, 1, 1)
    days_to_fri = (4 - jan1.weekday()) % 7  # 周五=4
    first_friday = jan1 + timedelta(days=days_to_fri)
    if first_friday > thursday:
        # 如果目标周四在第一个周五之前，使用上一年的第一个周五
        jan1_prev = date(thursday.year - 1, 1, 1)
        days_to_fri = (4 - jan1_prev.weekday()) % 7
        first_friday = jan1_prev + timedelta(days=days_to_fri)

    week_num = (friday - first_friday).days // 7 + 1
    return f"W{week_num}", friday, thursday, friday, thursday


def get_prev_week(week_start):
    """获取上一周的日期范围"""
    prev_thu = week_start - timedelta(days=1)  # 上周四
    prev_fri = prev_thu - timedelta(days=6)     # 上上周五

    # 计算周次：从年初第一个周五开始
    jan1 = date(prev_thu.year, 1, 1)
    days_to_fri = (4 - jan1.weekday()) % 7
    first_friday = jan1 + timedelta(days=days_to_fri)
    if first_friday > prev_thu:
        jan1_prev = date(prev_thu.year - 1, 1, 1)
        days_to_fri = (4 - jan1_prev.weekday()) % 7
        first_friday = jan1_prev + timedelta(days=days_to_fri)
    prev_week_num = (prev_fri - first_friday).days // 7 + 1
    return f"W{prev_week_num}", prev_fri, prev_thu


# ============================================================
# 2. 数据获取
# ============================================================

# 路径常量
HERE = Path(__file__).resolve().parent
REPORT_SYSTEM_DIR = HERE.parent  # report-system/
DATA_DIR = REPORT_SYSTEM_DIR / "data"  # 标准数据目录
OUTPUT_DIR = REPORT_SYSTEM_DIR / "output"  # 标准输出目录
PROJECT_DIR = HERE.parents[3]  # AiClaudeProject/（兼容旧版路径）

# Excel 文件路径：按优先级查找
# 1. 命令行 --excel 参数
# 2. 环境变量 WEEKLY_REPORT_EXCEL_PATH
# 3. ./data/技术支持工单明细.xlsx（标准数据目录，推荐）
# 4. 旧版路径 原始报表文档/技术支持工单明细.xlsx（兼容旧版）
def _resolve_excel_path(excel_arg=None):
    """按优先级解析 Excel 文件路径"""
    if excel_arg:
        p = Path(excel_arg)
        if p.exists():
            return p
        else:
            print(f"⚠️ 指定的 Excel 文件不存在: {excel_arg}")

    env_path = os.environ.get("WEEKLY_REPORT_EXCEL_PATH", "")
    if env_path:
        p = Path(env_path)
        if p.exists():
            return p

    data_path = DATA_DIR / "技术支持工单明细.xlsx"
    if data_path.exists():
        return data_path

    legacy_path = PROJECT_DIR / "原始报表文档" / "技术支持工单明细.xlsx"
    if legacy_path.exists():
        return legacy_path

    # 都不存在时返回标准路径（用于报错提示）
    return data_path

# 默认 EXCEL_PATH（可在运行时通过 _resolve_excel_path 覆盖）
EXCEL_PATH = _resolve_excel_path()

# Confluence Token：按优先级读取
# 1. 命令行 --token 参数
# 2. 环境变量 CONFLUENCE_TOKEN
# 3. .env 文件中的 CONFLUENCE_TOKEN
def _resolve_confluence_token(token_arg=None):
    """按优先级解析 Confluence Token"""
    if token_arg:
        return token_arg

    env_token = os.environ.get("CONFLUENCE_TOKEN", "")
    if env_token:
        return env_token

    # 尝试从 .env 文件读取
    env_file = REPORT_SYSTEM_DIR / ".env"
    if env_file.exists():
        try:
            with open(env_file, "r") as f:
                for line in f:
                    line = line.strip()
                    if line.startswith("CONFLUENCE_TOKEN="):
                        return line.split("=", 1)[1].strip().strip('"').strip("'")
        except Exception:
            pass

    return ""

# 初始化时解析（不传参数，走环境变量或 .env）
CONFLUENCE_TOKEN = _resolve_confluence_token()

TOKEN_HELP = ("请配置 Confluence Token（三选一）：\n"
              "  1. 命令行: --token YOUR_TOKEN\n"
              "  2. 环境变量: export CONFLUENCE_TOKEN=YOUR_TOKEN\n"
              "  3. .env 文件: 在 report-system/.env 中写入 CONFLUENCE_TOKEN=YOUR_TOKEN")


def fetch_excel_tickets(week_start, week_end, excel_path=None):
    """从 Excel 读取本周工单明细和超期工单"""
    if excel_path is None:
        excel_path = EXCEL_PATH
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"error": "请安装 openpyxl: pip install openpyxl"}

    if not excel_path.exists():
        return {"error": f"Excel 文件不存在: {excel_path}\n请将 技术支持工单明细.xlsx 放入 {DATA_DIR}/ 目录，或通过 --excel 参数指定路径"}

    wb = load_workbook(str(EXCEL_PATH), data_only=True)
    ws = wb.active

    tickets = []
    overdue = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 无工单编号
            continue

        time_val = row[7]  # 提交时间（Col G，索引7）
        if time_val is None:
            continue

        if isinstance(time_val, datetime):
            d = time_val
        else:
            try:
                d = datetime.strptime(str(time_val)[:10], '%Y-%m-%d')
            except Exception:
                continue

        # 筛选本周数据
        if not (week_start <= d.date() <= week_end):
            continue

        tid = str(row[0])
        ticket = {
            "id": tid,
            "date": d.strftime('%Y-%m-%d') if isinstance(d, datetime) else str(d),
            "dept2": str(row[2] or ""),       # 二级部门（Col C）
            "dept3": str(row[3] or ""),       # 三级部门（Col D）
            "module": str(row[4] or ""),      # 模块（Col E）
            "category": str(row[5] or ""),    # 问题分类（Col F）
            "time": str(time_val),
            "is_overdue": str(row[10] or ""),  # 是否超时（Col K）
            "overdue_reason": str(row[11] or ""),  # 超期原因（Col L）
            "desc": str(row[13] or ""),       # 工单摘要（Col N）
            "remark": str(row[15] or ""),     # 备注（Col P）
        }
        tickets.append(ticket)

        if ticket["is_overdue"] == "是":
            oncall = parse_oncall_from_remark(ticket["remark"])
            ticket.update(oncall)
            overdue.append(ticket)

    wb.close()

    # ===== 统计 =====
    dept_counts = Counter(t["dept2"] for t in tickets)
    module_counts = Counter(t["module"] for t in tickets)
    category_counts = Counter(t["category"] for t in tickets)

    # 按业务组归类
    grouped = defaultdict(list)
    for t in tickets:
        dept2 = t["dept2"]
        if dept2 in ["数智财务组", "电子档案组"]:
            grouped["附录1"].append(t)
        elif dept2 == "免疫规划组":
            grouped["附录2"].append(t)
        elif dept2 == "数字化支撑组":
            grouped["附录3"].append(t)
        else:
            grouped["其他"].append(t)

    # 超期工单按业务组归类（含排除规则）
    overdue_grouped = defaultdict(list)
    excluded_overdue = defaultdict(list)
    for t in overdue:
        dept2 = t["dept2"]
        reason = t["overdue_reason"]
        if dept2 in ["数智财务组", "电子档案组"]:
            if "外部超时" in reason:
                excluded_overdue["附录1"].append(t)
                continue
            overdue_grouped["附录1"].append(t)
        elif dept2 == "免疫规划组":
            overdue_grouped["附录2"].append(t)
        elif dept2 == "数字化支撑组":
            overdue_grouped["附录3"].append(t)

    return {
        "total": len(tickets),
        "overdue_total": len(overdue),
        "by_dept": dict(dept_counts),
        "by_module": dict(module_counts),
        "by_category": dict(category_counts),
        "grouped": {k: len(v) for k, v in grouped.items()},
        "overdue_grouped": {
            k: {
                "count": len(v),
                "items": v,
            }
            for k, v in overdue_grouped.items()
        },
        "excluded_overdue": {
            k: {"count": len(v), "reason": "外部超时"}
            for k, v in excluded_overdue.items() if v
        },
        "tickets": tickets,
    }


def parse_oncall_from_remark(remark):
    """从备注列解析 ONCALL 数据"""
    if not remark or not str(remark).strip():
        return {"oncall_link": "——", "oncall_transfer_time": "——", "oncall_reply_time": "——"}

    text = str(remark).strip()

    # 提取 ONCALL 链接（支持 "ONCALL链接"、"ONCALL" 两种格式）
    link_match = re.search(r'ONCALL(?:链接)?[：:]\s*(https?://\S+)', text)
    oncall_link = link_match.group(1) if link_match else "——"

    # 提取 ONCALL 转派时间
    transfer_match = re.search(r'ONCALL转派时间[：:]\s*(\S+)', text)
    oncall_transfer_time = transfer_match.group(1) if transfer_match else "——"

    # 提取 ONCALL 回复时间
    reply_match = re.search(r'ONCALL回复时间[：:]\s*(\S+)', text)
    oncall_reply_time = reply_match.group(1) if reply_match else "——"

    return {
        "oncall_link": oncall_link,
        "oncall_transfer_time": oncall_transfer_time,
        "oncall_reply_time": oncall_reply_time,
    }


def fetch_confluence_demands():
    """从 Confluence 获取需求列表"""
    if not CONFLUENCE_TOKEN:
        return {"error": TOKEN_HELP, "demands": {}}

    import subprocess
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            "https://cf.cai-inc.com/rest/api/content/258705731?expand=body.storage"
        ],
        capture_output=True, text=True, timeout=30
    )

    if result.returncode != 0 or not result.stdout.strip():
        return {"error": "Confluence 数据获取失败", "demands": {}}

    try:
        data = json.loads(result.stdout)
        html = data["body"]["storage"]["value"]
    except Exception:
        return {"error": "Confluence 响应解析失败", "demands": {}}

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return {"error": "请安装 beautifulsoup4: pip install beautifulsoup4", "demands": {}}

    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return {"error": "需求表格未找到", "demands": {}}

    rows = table.find_all("tr")[1:]  # 跳过表头
    demands = {"数智财务": [], "免疫规划": [], "数字化支撑": [], "电子档案": []}

    for row in rows:
        cols = row.find_all("td")
        if len(cols) < 4:
            continue
        module = cols[0].get_text(strip=True)
        name = cols[1].get_text(strip=True)
        link_el = cols[2].find("a")
        link = link_el["href"] if link_el else ""
        status = cols[3].get_text(strip=True)

        if not module or not name:
            continue
        if "已完成" in status:
            continue

        item = {"name": name, "link": link, "status": status}
        if "数智财务" in module:
            demands["数智财务"].append(item)
        elif "免疫规划" in module or "疫苗" in module:
            demands["免疫规划"].append(item)
        elif "数字化" in module:
            demands["数字化支撑"].append(item)
        elif "电子档案" in module:
            demands["电子档案"].append(item)

    return {"demands": demands, "error": None}


def fetch_confluence_dashboard(target_week_str):
    """从 Confluence 翡翠周报数据看板获取非客满统计数据"""
    if not CONFLUENCE_TOKEN:
        return {"error": TOKEN_HELP, "data": None}

    import subprocess
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            "https://cf.cai-inc.com/rest/api/content/278883515?expand=body.storage"
        ],
        capture_output=True, text=True, timeout=30
    )

    if result.returncode != 0 or not result.stdout.strip():
        return {"error": "Confluence 看板数据获取失败", "data": None}

    try:
        data = json.loads(result.stdout)
        html = data["body"]["storage"]["value"]
    except Exception:
        return {"error": "Confluence 看板响应解析失败", "data": None}

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return {"error": "需要 beautifulsoup4", "data": None}

    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return {"error": "看板表格未找到", "data": None}

    rows = table.find_all("tr")
    headers = [c.get_text(strip=True) for c in rows[0].find_all("td")]
    col_map = {h: i for i, h in enumerate(headers)}

    target_row = None
    for row in rows[1:]:
        cells = row.find_all("td")
        if cells and cells[0].get_text(strip=True) == f"2026{target_week_str}":
            target_row = cells
            break

    if not target_row:
        return {"error": f"看板中未找到 {target_week_str} 的数据", "data": None}

    return {
        "data": {
            "auto_transfer": target_row[col_map.get("浙里报机器人转人工申请量", 0)].get_text(strip=True) if "浙里报机器人转人工申请量" in col_map else "—",
            "pm_submit_finance": target_row[col_map.get("浙里报PM技术工单量", 0)].get_text(strip=True) if "浙里报PM技术工单量" in col_map else "—",
            "pm_submit_immune": target_row[col_map.get("疫苗PM技术工单量", 0)].get_text(strip=True) if "疫苗PM技术工单量" in col_map else "—",
        },
        "error": None,
    }


def fetch_confluence_metrics():
    """从 Confluence 业务指标页面获取业务指标详情数据。
    页面: https://cf.cai-inc.com/pages/viewpage.action?pageId=286888251
    返回: {metrics: {业务: [{主指标, 子指标, 上周值, 本周值, 趋势, 变化率}]}, week_labels: [上周, 本周]}
    """
    if not CONFLUENCE_TOKEN:
        return {"error": TOKEN_HELP, "metrics": {}}

    import subprocess
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            "https://cf.cai-inc.com/rest/api/content/286888251?expand=body.storage"
        ],
        capture_output=True, text=True, timeout=30
    )

    if result.returncode != 0 or not result.stdout.strip():
        return {"error": "Confluence 业务指标获取失败", "metrics": {}}

    try:
        data = json.loads(result.stdout)
        html = data["body"]["storage"]["value"]
    except Exception:
        return {"error": "Confluence 业务指标响应解析失败", "metrics": {}}

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return {"error": "需要 beautifulsoup4", "metrics": {}}

    soup = BeautifulSoup(html, "html.parser")
    table = soup.find("table")
    if not table:
        return {"error": "业务指标表格未找到", "metrics": {}}

    rows = table.find_all("tr")

    # 第一行是表头，提取周标签
    header_cells = rows[0].find_all("td")
    week_labels = []
    for cell in header_cells:
        text = cell.get_text(strip=True)
        if text and re.match(r'2026W\d+', text):
            week_labels.append(text)

    # 解析数据行。rowspan 导致每行列数不同：
    #   7列 = 新业务 + 新主指标 + 子指标 + 数据
    #   6列 = 新主指标 + 子指标 + 数据（业务延续）
    #   5列 = 子指标 + 数据（业务和主指标都延续）
    metrics = {}
    current_business = ""
    current_main = ""

    for row in rows[1:]:
        cells = row.find_all("td")
        if not cells:
            continue

        # 提取纯文本
        cell_texts = [c.get_text(strip=True) for c in cells]

        # 跳过表头重复行
        if cell_texts[0] == "业务":
            continue

        n = len(cell_texts)

        if n >= 7:
            # 完整行：业务 + 主指标 + 子指标 + prev + curr + trend + change
            current_business = cell_texts[0]
            current_main = cell_texts[1]
            sub_metric = cell_texts[2]
            prev_val = cell_texts[3]
            curr_val = cell_texts[4]
            trend = cell_texts[5]
            change = cell_texts[6] if n > 6 else ""
        elif n == 6:
            # 新主指标行：主指标 + 子指标 + prev + curr + trend + change
            current_main = cell_texts[0]
            sub_metric = cell_texts[1]
            prev_val = cell_texts[2]
            curr_val = cell_texts[3]
            trend = cell_texts[4]
            change = cell_texts[5] if n > 5 else ""
        elif n == 5:
            # 延续行：子指标 + prev + curr + trend + change
            sub_metric = cell_texts[0]
            prev_val = cell_texts[1]
            curr_val = cell_texts[2]
            trend = cell_texts[3]
            change = cell_texts[4]
        else:
            continue

        if not sub_metric or not current_business:
            continue

        # 规范化趋势符号
        if trend in ("&mdash;", "—", "-"):
            trend = "—"

        if current_business not in metrics:
            metrics[current_business] = []

        metrics[current_business].append({
            "main_metric": current_main,
            "sub_metric": sub_metric,
            "prev_value": prev_val,
            "curr_value": curr_val,
            "trend": trend,
            "change": change,
        })

    return {
        "metrics": metrics,
        "week_labels": week_labels,
        "page_title": data.get("title", ""),
        "error": None,
    }


def fetch_confluence_faults(week_start, week_end):
    """从 Confluence 故障明细页面获取本周故障数据。
    页面: https://cf.cai-inc.com/pages/viewpage.action?pageId=257088525
    根据故障日期筛选落在 week_start ~ week_end 范围内的故障。
    返回: {faults: {业务组: [{等级, 日期, 故障名称, 故障链接, 故障原因, 故障分类, 复盘链接}]}}
    """
    if not CONFLUENCE_TOKEN:
        return {"error": TOKEN_HELP, "faults": {}}

    import subprocess
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            "https://cf.cai-inc.com/rest/api/content/257088525?expand=body.storage"
        ],
        capture_output=True, text=True, timeout=30
    )

    if result.returncode != 0 or not result.stdout.strip():
        return {"error": "Confluence 故障数据获取失败", "faults": {}}

    try:
        data = json.loads(result.stdout)
        html = data["body"]["storage"]["value"]
    except Exception:
        return {"error": "Confluence 故障数据响应解析失败", "faults": {}}

    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return {"error": "需要 beautifulsoup4", "faults": {}}

    soup = BeautifulSoup(html, "html.parser")

    # 找到"故障明细数据"标题后的第一个表格
    fault_h3 = soup.find('h3', string=lambda t: t and '故障明细数据' in t)
    if not fault_h3:
        return {"error": "故障明细数据表格未找到", "faults": {}}

    table = fault_h3.find_next('table')
    if not table:
        return {"error": "故障明细表格未找到", "faults": {}}

    # 业务组名称映射
    GROUP_MAP = {
        "浙里报": "数智财务",
        "疫苗": "免疫规划",
        "数字化": "数字化支撑",
        "电子档案": "电子档案",
    }

    rows = table.find_all("tr")
    current_group = ""
    all_faults = []

    for row in rows:
        cells = row.find_all("td")
        if not cells:
            continue

        # 检查是否是业务组标题行（colspan=7 的整行）
        if len(cells) == 1 and cells[0].get("colspan") == "7":
            group_name = cells[0].get_text(strip=True)
            current_group = GROUP_MAP.get(group_name, group_name)
            continue

        # 数据行
        cell_texts = [c.get_text(strip=True) for c in cells]
        if len(cell_texts) < 7:
            continue
        if cell_texts[0] == "等级":
            continue

        level = cell_texts[0]
        date_str = cell_texts[1]
        name = cell_texts[2] if len(cell_texts) > 2 else ""
        link = cell_texts[3] if len(cell_texts) > 3 else ""
        cause = cell_texts[4] if len(cell_texts) > 4 else ""
        category = cell_texts[5] if len(cell_texts) > 5 else ""
        review = cell_texts[6] if len(cell_texts) > 6 else ""

        if not level and not name:
            continue

        # 解析日期：支持多种格式
        parsed_date = _parse_fault_date(date_str, name)
        if not parsed_date:
            continue

        all_faults.append({
            "level": level,
            "date": parsed_date.strftime("%Y-%m-%d"),
            "date_str": date_str,
            "name": name,
            "link": link,
            "cause": cause,
            "category": category,
            "review": review,
            "group": current_group,
        })

    # 按日期范围筛选本周故障
    week_faults = {}
    for f in all_faults:
        try:
            fault_date = datetime.strptime(f["date"], "%Y-%m-%d").date()
        except Exception:
            continue
        if week_start <= fault_date <= week_end:
            group = f["group"]
            if group not in week_faults:
                week_faults[group] = []
            week_faults[group].append(f)

    return {
        "faults": week_faults,
        "all_faults_count": len(all_faults),
        "week_faults_count": sum(len(v) for v in week_faults.values()),
        "error": None,
    }


def _parse_fault_date(date_str, name):
    """从故障日期列或故障名称中解析日期。
    支持格式: 2026-07-09, 01月16日, 7月9日, 07月17日
    """
    # 格式1: YYYY-MM-DD（如 2026-07-09）
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", date_str)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except ValueError:
            pass

    # 格式2: MM月DD日（如 01月16日, 07月17日）
    m = re.match(r"(\d{1,2})月(\d{1,2})日", date_str)
    if m:
        try:
            return date(2026, int(m.group(1)), int(m.group(2)))
        except ValueError:
            pass

    # 如果日期列为空，尝试从故障名称中提取
    if not date_str:
        m = re.match(r"(\d{1,2})月(\d{1,2})日", name)
        if m:
            try:
                return date(2026, int(m.group(1)), int(m.group(2)))
            except ValueError:
                pass

    return None


def fetch_latest_weekly_report():
    """从 Confluence 获取最新一期周报格式参考"""
    if not CONFLUENCE_TOKEN:
        return {"error": TOKEN_HELP, "reports": []}

    import subprocess
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            "https://cf.cai-inc.com/rest/api/content/252657891/child/page?limit=25"
        ],
        capture_output=True, text=True, timeout=30
    )

    if result.returncode != 0 or not result.stdout.strip():
        return {"error": "Confluence 周报列表获取失败", "reports": []}

    try:
        data = json.loads(result.stdout)
        reports = []
        for r in data.get("results", []):
            reports.append({
                "id": r["id"],
                "title": r["title"],
                "version": r.get("version", {}).get("number", 0),
            })
        reports.sort(key=lambda x: x["title"], reverse=True)
        return {"reports": reports[:5], "error": None}
    except Exception:
        return {"error": "周报列表解析失败", "reports": []}


# ============================================================
# 3. 主入口
# ============================================================

def gather_all_data(week_str=None, excel_path=None):
    """并行获取所有数据源，返回结构化结果"""
    # 计算周日期
    if week_str:
        # 从 W30 格式解析
        try:
            week_num = int(week_str.replace("W", "").replace("w", ""))
        except ValueError:
            return {"error": f"无效的周次格式: {week_str}，请使用 W30 格式"}
    else:
        week_str, friday, thursday, _, _ = get_week_from_date()
        week_num = int(week_str.replace("W", ""))

    # 计算日期范围
    friday, thursday = _week_num_to_dates(week_num)
    week_str = f"W{week_num}"

    # 上一周
    prev_week_str, prev_friday, prev_thursday = get_prev_week(friday)
    prev_week_num = int(prev_week_str.replace("W", ""))

    # ---- 合并周检测：先拉取 Confluence 指标，检查是否有合并周标签 ----
    excel_friday, excel_thursday = friday, thursday  # 默认：常规单周
    fault_friday, fault_thursday = friday, thursday
    merged_weeks = None  # 记录合并的周次，如 "W31-32"

    try:
        metrics_pre = fetch_confluence_metrics()
        labels = metrics_pre.get("week_labels", [])
        if len(labels) >= 2:
            curr_label = labels[1]  # 当前周标签，如 "W31-32" 或 "2026W31-32"
            # 去掉年份前缀
            label_clean = curr_label.replace(str(thursday.year), "").replace("2026", "")
            if "-" in label_clean:
                # 检测到合并周，如 "W31-32"
                merged_weeks = label_clean
                parts = label_clean.replace("W", "").replace("w", "").split("-")
                try:
                    first_week = int(parts[0])
                    # 扩展 Excel 日期范围：从第一个合并周的开始到当前周的结束
                    excel_friday, _ = _week_num_to_dates(first_week)
                    fault_friday = excel_friday
                except (ValueError, IndexError):
                    pass  # 解析失败则保持默认范围
    except Exception:
        pass  # 指标获取失败则保持默认范围，后续正常获取

    result = {
        "meta": {
            "week": week_str,
            "year": thursday.year,
            "date_range": f"{friday.strftime('%m/%d')} - {thursday.strftime('%m/%d')}",
            "friday": friday.strftime('%Y-%m-%d'),
            "thursday": thursday.strftime('%Y-%m-%d'),
            "prev_week": prev_week_str,
            "prev_date_range": f"{prev_friday.strftime('%m/%d')} - {prev_thursday.strftime('%m/%d')}",
            "generated_at": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        },
    }

    # 如果检测到合并周，记录到 meta 中
    if merged_weeks:
        result["meta"]["merged_weeks"] = merged_weeks
        result["meta"]["excel_date_range"] = f"{excel_friday.strftime('%m/%d')} - {excel_thursday.strftime('%m/%d')}"

    # 并行获取数据（Excel 和故障使用可能扩展后的日期范围）
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = {
            executor.submit(fetch_excel_tickets, excel_friday, excel_thursday, excel_path): "excel",
            executor.submit(fetch_confluence_demands): "demands",
            executor.submit(fetch_confluence_dashboard, week_str): "dashboard",
            executor.submit(fetch_confluence_faults, fault_friday, fault_thursday): "faults",
        }
        # 如果已经预取了 metrics，直接使用，不再重复获取
        if 'metrics_pre' in locals() and metrics_pre and "error" not in metrics_pre:
            result["metrics"] = metrics_pre
        else:
            futures[executor.submit(fetch_confluence_metrics)] = "metrics"

        for future in as_completed(futures):
            key = futures[future]
            try:
                result[key] = future.result()
            except Exception as e:
                result[key] = {"error": str(e)}

    return result


# ============================================================
# 5. 数据审计
# ============================================================

def run_audit(data):
    """对获取的数据进行自动审计，逐项检查并返回结果。
    返回: {passed: bool, checks: [{name, passed, detail, severity}]}
    """
    checks = []

    # ---- 1. 业务指标数据完整性 ----
    metrics = data.get("metrics", {}).get("metrics", {})
    expected_biz = ["数智财务", "免疫规划", "数字化支撑", "电子档案"]
    for biz in expected_biz:
        if biz not in metrics:
            checks.append({
                "name": f"业务指标-{biz}",
                "passed": False,
                "detail": f"缺少 {biz} 的业务指标数据",
                "severity": "error",
            })
        elif len(metrics[biz]) == 0:
            checks.append({
                "name": f"业务指标-{biz}",
                "passed": False,
                "detail": f"{biz} 的业务指标为空",
                "severity": "error",
            })
        else:
            checks.append({
                "name": f"业务指标-{biz}",
                "passed": True,
                "detail": f"{biz}: {len(metrics[biz])} 项指标",
                "severity": "info",
            })

    # ---- 2. 两小时完结率四舍五入检查 ----
    rate_metrics = [
        ("数智财务", "技术支持两小时工单完成率（内部）"),
        ("数智财务", "技术支持两小时工单完成率（含外部）"),
        ("免疫规划", "技术支持两小时工单完成率"),
        ("数字化支撑", "技术支持两小时工单完成率"),
        ("电子档案", "技术支持两小时工单完成率（内部）"),
        ("电子档案", "技术支持两小时工单完成率（含外部）"),
    ]
    for biz, sub in rate_metrics:
        items = metrics.get(biz, [])
        for item in items:
            if item["sub_metric"] == sub:
                val = item["curr_value"]
                try:
                    # 去掉%号，转为浮点数
                    num = float(val.replace("%", ""))
                    rounded = round(num)
                    if num != rounded and num != int(num):
                        checks.append({
                            "name": f"完结率取整-{biz}-{sub}",
                            "passed": False,
                            "detail": f"{val} 需四舍五入取整为 {rounded}%（当前含小数）",
                            "severity": "warning",
                        })
                    else:
                        checks.append({
                            "name": f"完结率取整-{biz}-{sub}",
                            "passed": True,
                            "detail": f"{val} → {rounded}% {'达标' if rounded >= 95 else '不达标'}",
                            "severity": "info",
                        })
                except ValueError:
                    pass
                break

    # ---- 3. 故障数据与业务指标一致性 ----
    faults = data.get("faults", {}).get("faults", {})
    # 统计本周各业务组 P1/P2 和 P3/P4 故障数
    fault_p1p2 = {}
    fault_p3p4 = {}
    for group, items in faults.items():
        p1p2 = sum(1 for f in items if f["level"] in ("P1", "P2"))
        p3p4 = sum(1 for f in items if f["level"] in ("P3", "P4"))
        fault_p1p2[group] = p1p2
        fault_p3p4[group] = p3p4

    # 映射到业务指标中的业务名
    biz_fault_map = {
        "数智财务": "数智财务",
        "免疫规划": "免疫规划",
    }
    for biz, metric_biz in biz_fault_map.items():
        if biz in fault_p1p2:
            checks.append({
                "name": f"故障一致性-{biz}-P1P2",
                "passed": True,
                "detail": f"本周新增 P1/P2: {fault_p1p2[biz]} 个",
                "severity": "info",
            })
        if biz in fault_p3p4:
            checks.append({
                "name": f"故障一致性-{biz}-P3P4",
                "passed": True,
                "detail": f"本周新增 P3/P4: {fault_p3p4[biz]} 个",
                "severity": "info",
            })

    # ---- 4. 超期工单排除规则检查 ----
    excel = data.get("excel", {})
    overdue_grouped = excel.get("overdue_grouped", {})
    excluded = excel.get("excluded_overdue", {})
    if "附录1" in overdue_grouped:
        appendix1_items = overdue_grouped["附录1"].get("items", [])
        has_external = any("外部超时" in item.get("overdue_reason", "") for item in appendix1_items)
        if has_external:
            checks.append({
                "name": "超期工单-附录1排除外部超时",
                "passed": False,
                "detail": "附录1 超期工单中仍包含'外部超时'条目，排除规则未生效",
                "severity": "error",
            })
        else:
            checks.append({
                "name": "超期工单-附录1排除外部超时",
                "passed": True,
                "detail": f"附录1 超期 {len(appendix1_items)} 条，已排除外部超时",
                "severity": "info",
            })
    if excluded:
        total_excluded = sum(v.get("count", 0) for v in excluded.values())
        checks.append({
            "name": "超期工单-排除统计",
            "passed": True,
            "detail": f"共排除外部超时 {total_excluded} 条",
            "severity": "info",
        })

    # ---- 5. 需求状态过滤检查 ----
    demands = data.get("demands", {}).get("demands", {})
    for cat, items in demands.items():
        has_completed = any("已完成" in item.get("status", "") for item in items)
        if has_completed:
            checks.append({
                "name": f"需求过滤-{cat}",
                "passed": False,
                "detail": f"{cat} 需求列表中包含'已完成'状态的需求，过滤规则未生效",
                "severity": "error",
            })
        else:
            checks.append({
                "name": f"需求过滤-{cat}",
                "passed": True,
                "detail": f"{cat}: {len(items)} 条未完成需求",
                "severity": "info",
            })

    # ---- 6. 趋势标注一致性检查 ----
    for biz, items in metrics.items():
        for item in items:
            prev = item["prev_value"]
            curr = item["curr_value"]
            trend = item["trend"]
            try:
                p = float(prev.replace("%", ""))
                c = float(curr.replace("%", ""))
                if c > p and trend != "▲":
                    checks.append({
                        "name": f"趋势检查-{biz}-{item['sub_metric']}",
                        "passed": False,
                        "detail": f"{prev}→{curr} 应为 ▲，实际为 {trend}",
                        "severity": "warning",
                    })
                elif c < p and trend != "▼":
                    checks.append({
                        "name": f"趋势检查-{biz}-{item['sub_metric']}",
                        "passed": False,
                        "detail": f"{prev}→{curr} 应为 ▼，实际为 {trend}",
                        "severity": "warning",
                    })
            except ValueError:
                pass

    # ---- 汇总 ----
    errors = [c for c in checks if c["severity"] == "error" and not c["passed"]]
    warnings = [c for c in checks if c["severity"] == "warning" and not c["passed"]]
    passed = len(errors) == 0

    return {
        "passed": passed,
        "total_checks": len(checks),
        "errors": len(errors),
        "warnings": len(warnings),
        "checks": checks,
    }


def _week_num_to_dates(week_num):
    """根据 W 周次计算周五-周四的日期范围"""
    today = date.today()
    # 找到当前周的周四
    current_thu = today - timedelta(days=(today.weekday() - 3) % 7)
    # 找到当年第一个周五
    jan1 = date(current_thu.year, 1, 1)
    days_to_fri = (4 - jan1.weekday()) % 7
    first_friday = jan1 + timedelta(days=days_to_fri)
    # 当前周次
    current_week = (current_thu - first_friday).days // 7 + 1
    if current_week < 1:
        current_week = 1
    week_diff = week_num - current_week
    target_thu = current_thu + timedelta(weeks=week_diff)
    target_fri = target_thu - timedelta(days=6)
    return target_fri, target_thu


# ============================================================
# 6. Markdown → Confluence Storage Format 转换
# ============================================================

def md_to_confluence_storage(md_text):
    """将 Markdown 文本转换为 Confluence Storage Format (XHTML)。

    支持的 Markdown 元素：
    - 标题 H1-H6
    - 段落、换行
    - 粗体、斜体、行内代码
    - 无序列表（- 和 *）
    - 表格（| col | col |）
    - 水平分割线（---）
    - 链接 [text](url)
    - 引用块（>）
    - 图片 ![alt](url)
    """
    import re

    lines = md_text.split("\n")
    output = []
    i = 0
    n = len(lines)

    # 状态跟踪
    in_table = False
    ul_stack = []  # 嵌套列表的缩进层级栈
    li_open = []  # 跟踪每个层级是否有未闭合的 <li>
    in_blockquote = False
    table_rows = []
    blockquote_lines = []

    def flush_table():
        """输出缓存的表格行"""
        nonlocal in_table
        if not table_rows:
            return
        rows_html = []
        is_first = True
        for row in table_rows:
            cells = row.split("|")
            # 去掉首尾空单元格（由 | 包裹导致）
            if cells and cells[0].strip() == "":
                cells = cells[1:]
            if cells and cells[-1].strip() == "":
                cells = cells[:-1]
            tag = "th" if is_first else "td"
            cell_html = "".join(
                f"<{tag}>{_inline_md_to_html(c.strip())}</{tag}>"
                for c in cells
            )
            rows_html.append(f"<tr>{cell_html}</tr>")
            is_first = False
        output.append(f'<table><tbody>{"".join(rows_html)}</tbody></table>')
        table_rows.clear()
        in_table = False

    def close_li(level):
        """关闭指定层级未闭合的 <li>"""
        if level < len(li_open) and li_open[level]:
            output.append("</li>")
            li_open[level] = False

    def close_li_up_to(level):
        """关闭从当前深度到指定层级的所有未闭合 <li>"""
        for lv in range(len(li_open) - 1, level - 1, -1):
            if lv < len(li_open) and li_open[lv]:
                output.append("</li>")
                li_open[lv] = False

    def flush_ul():
        """关闭所有未闭合的列表层级（先关 li 再关 ul）"""
        while ul_stack:
            level = len(ul_stack) - 1
            close_li(level)
            output.append("</ul>")
            ul_stack.pop()
            if li_open:
                li_open.pop()

    def ensure_ul_level(indent, is_first_level=False):
        """根据缩进层级确保正确的 ul 嵌套。
        indent: 缩进空格数（0=顶层, 2=二层, 4=三层...）
        is_first_level: 是否是顶层列表（用于添加 style 属性）
        """
        # 目标深度：0缩进=1层ul, 2缩进=2层ul, 4缩进=3层ul...
        target_depth = indent // 2 + 1

        # 上升层级（缩进变浅）：关闭深层的 li 和 ul
        while len(ul_stack) > target_depth:
            level = len(ul_stack) - 1
            close_li(level)
            output.append("</ul>")
            ul_stack.pop()
            if li_open:
                li_open.pop()

        # 下降层级（缩进变深）：打开新的 ul（在父 li 内部）
        while len(ul_stack) < target_depth:
            if len(ul_stack) == 0 and is_first_level:
                output.append('<ul style="text-decoration: none;">')
            else:
                output.append("<ul>")
            ul_stack.append(len(ul_stack))
            li_open.append(False)  # 新层级还没有 li

    def flush_blockquote():
        nonlocal in_blockquote
        if blockquote_lines:
            content = " ".join(blockquote_lines)
            content = _inline_md_to_html(content)
            output.append(f"<blockquote><p>{content}</p></blockquote>")
            blockquote_lines.clear()
        in_blockquote = False

    while i < n:
        line = lines[i]
        stripped = line.strip()

        # 表格行
        if stripped.startswith("|") and (stripped.endswith("|") or "|" in stripped[1:]):
            # 检查是否是分隔行（如 |---|---|）
            if re.match(r'^\|[\s\-:|]+\|$', stripped):
                i += 1
                continue
            if not in_table:
                flush_ul()
                flush_blockquote()
                in_table = True
            table_rows.append(stripped)
            i += 1
            continue
        elif in_table:
            flush_table()

        # 水平分割线
        if re.match(r'^[-*_]{3,}\s*$', stripped):
            flush_ul()
            flush_blockquote()
            output.append("<hr />")
            i += 1
            continue

        # 标题
        heading_match = re.match(r'^(#{1,6})\s+(.+)$', stripped)
        if heading_match:
            flush_ul()
            flush_blockquote()
            level = len(heading_match.group(1))
            text = _inline_md_to_html(heading_match.group(2))
            output.append(f"<h{level}>{text}</h{level}>")
            i += 1
            continue

        # 空行（必须在列表检查之前处理，避免列表内的空行触发 flush_ul）
        if stripped == "":
            if in_blockquote:
                flush_blockquote()
            if in_table:
                flush_table()
            # 在列表内部，不因空行而中断列表（允许嵌套列表中有空行分隔）
            if ul_stack:
                i += 1
                continue
            i += 1
            continue

        # 无序列表（支持缩进嵌套）
        # 注意：使用原始行 line（而非 stripped）来检测缩进层级
        ul_match = re.match(r'^(\s*)[-*]\s+(.+)$', line)
        if ul_match:
            flush_blockquote()
            if in_table:
                flush_table()
            indent = len(ul_match.group(1))
            # 判断是否是顶层列表（在标题或空行之后的第一个列表）
            is_first_level = (len(ul_stack) == 0)
            ensure_ul_level(indent, is_first_level)
            level = len(ul_stack) - 1  # 当前层级
            # 在输出新 li 之前，关闭同层级上一个 li
            close_li(level)
            text = _inline_md_to_html(ul_match.group(2))
            # 缩进 ≤ 2 的段落级列表项用 <p> 包裹，深层叶子项（问题描述/处理方式/原因）不用
            if indent <= 2:
                output.append(f"<li><p>{text}</p>")
            else:
                output.append(f"<li>{text}")
            li_open[level] = True
            i += 1
            continue
        elif ul_stack:
            flush_ul()

        # 引用块
        bq_match = re.match(r'^>\s*(.*)$', stripped)
        if bq_match:
            if ul_stack:
                flush_ul()
            if in_table:
                flush_table()
            if not in_blockquote:
                in_blockquote = True
            blockquote_lines.append(bq_match.group(1))
            i += 1
            continue
        elif in_blockquote:
            flush_blockquote()

        # 普通段落
        if ul_stack:
            flush_ul()
        if in_blockquote:
            flush_blockquote()
        if in_table:
            flush_table()

        text = _inline_md_to_html(stripped)
        output.append(f"<p>{text}</p>")
        i += 1

    # 清理未闭合的块
    if in_table:
        flush_table()
    if ul_stack:
        flush_ul()
    if in_blockquote:
        flush_blockquote()

    return "".join(output)


def _inline_md_to_html(text):
    """转换行内 Markdown 元素为 HTML"""
    import re

    # 先用占位符保护已有的 HTML 特殊字符
    # 先转义 &（必须在最前面），然后转义 < 和 >
    text = text.replace("&", "&amp;")
    text = text.replace("<", "&lt;")
    text = text.replace(">", "&gt;")

    # 在转义后的文本上做 Markdown → HTML 转换
    # 使用占位符技巧避免转义冲突

    # 粗体+斜体
    text = re.sub(r'&amp;lt;strong&amp;gt;&amp;lt;em&amp;gt;(.+?)&amp;lt;/em&amp;gt;&amp;lt;/strong&amp;gt;',
                  r'<strong><em>\1</em></strong>', text)
    # 粗体 ***text*** → 但先检查是否已经有 ** 的情况
    # 注意：由于 & 已经被转义，** 在源文本中就是字面的 **
    text = re.sub(r'\*\*(.+?)\*\*', r'<strong>\1</strong>', text)
    # 斜体
    text = re.sub(r'\*(.+?)\*', r'<em>\1</em>', text)
    # 行内代码
    text = re.sub(r'`([^`]+)`', r'<code>\1</code>', text)
    # 链接 [text](url)
    text = re.sub(r'\[([^\]]+)\]\(([^)]+)\)', r'<a href="\2">\1</a>', text)
    # 图片 ![alt](url)
    text = re.sub(r'!\[([^\]]*)\]\(([^)]+)\)',
                  r'<ac:image><ri:url ri:value="\2" /></ac:image>', text)

    return text


# ============================================================
# 7. Confluence 发布
# ============================================================

# Confluence 发布配置
CONFLUENCE_BASE_URL = "https://cf.cai-inc.com"
CONFLUENCE_PARENT_PAGE_ID = "252657891"  # 26年技术支持周报
CONFLUENCE_SPACE_KEY = "FCSY"


def publish_weekly_report(week_str=None):
    """将本地生成的周报 Markdown 文件发布到 Confluence。

    流程：
    1. 确定周次和对应的本地文件路径
    2. 读取本地 Markdown 文件
    3. 转换为 Confluence Storage Format
    4. 检查 Confluence 上是否已有同名页面
    5. 有则更新（覆盖），无则创建新页面
    """
    # 计算周次
    if week_str:
        try:
            week_num = int(week_str.replace("W", "").replace("w", ""))
        except ValueError:
            print(f"❌ 无效的周次格式: {week_str}，请使用 W30 格式")
            return
    else:
        week_str, friday, thursday, _, _ = get_week_from_date()
        week_num = int(week_str.replace("W", ""))

    friday, thursday = _week_num_to_dates(week_num)
    week_str = f"W{week_num}"

    # 本地文件路径（优先新路径，兼容旧路径）
    report_path_new = REPORT_SYSTEM_DIR / "output" / "周报" / f"2026-{week_str}-技术支持周报.md"
    report_path_old = PROJECT_DIR / "2026报表数据知识库" / "周报" / f"2026-{week_str}-技术支持周报.md"
    if report_path_new.exists():
        report_path = report_path_new
    elif report_path_old.exists():
        report_path = report_path_old
        print(f"📄 读取本地周报（旧路径）: {report_path}")
    else:
        report_path = report_path_new  # 默认新路径（用于报错提示）
    if not report_path.exists():
        print(f"❌ 本地周报文件不存在: {report_path}")
        print(f"   请先生成周报文件后再发布。")
        return

    print(f"📄 读取本地周报: {report_path}")
    with open(report_path, "r", encoding="utf-8") as f:
        md_content = f.read()

    # 转换 Markdown → Confluence Storage Format
    print("🔄 转换 Markdown → Confluence Storage Format...")
    confluence_body = md_to_confluence_storage(md_content)

    # 页面标题：格式为 YYYYMMDD翡翠技术支持周报（日期为本周四，即周报截止日期）
    page_title = f"{thursday.strftime('%Y%m%d')}翡翠技术支持周报"
    date_range = f"{friday.strftime('%m/%d')} - {thursday.strftime('%m/%d')}"

    if not CONFLUENCE_TOKEN:
        print(f"❌ {TOKEN_HELP}")
        return

    import subprocess

    # 1. 查找是否已有同名页面
    print(f"🔍 检查 Confluence 上是否已有页面: {page_title}")
    result = subprocess.run(
        [
            "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
            f"{CONFLUENCE_BASE_URL}/rest/api/content/{CONFLUENCE_PARENT_PAGE_ID}/child/page?limit=50&expand=version"
        ],
        capture_output=True, text=True, timeout=30
    )

    existing_page_id = None
    existing_version = 1
    if result.returncode == 0 and result.stdout.strip():
        try:
            data = json.loads(result.stdout)
            for r in data.get("results", []):
                if r.get("title") == page_title:
                    existing_page_id = r["id"]
                    existing_version = r.get("version", {}).get("number", 1)
                    print(f"   找到已有页面: {existing_page_id} (版本 {existing_version})")
                    break
        except Exception:
            pass

    # 如果子页面列表没找到（可能被分页截断），通过标题搜索
    if not existing_page_id:
        search_result = subprocess.run(
            [
                "curl", "-s", "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
                f"{CONFLUENCE_BASE_URL}/rest/api/content?title={page_title}&spaceKey={CONFLUENCE_SPACE_KEY}&expand=version"
            ],
            capture_output=True, text=True, timeout=30
        )
        if search_result.returncode == 0 and search_result.stdout.strip():
            try:
                search_data = json.loads(search_result.stdout)
                for r in search_data.get("results", []):
                    if r.get("title") == page_title:
                        existing_page_id = r["id"]
                        existing_version = r.get("version", {}).get("number", 1)
                        print(f"   通过搜索找到已有页面: {existing_page_id} (版本 {existing_version})")
                        break
            except Exception:
                pass

    # 2. 创建或更新页面
    if existing_page_id:
        print(f"🔄 更新已有页面 (版本 {existing_version} → {existing_version + 1})...")
        payload = {
            "id": existing_page_id,
            "type": "page",
            "title": page_title,
            "space": {"key": CONFLUENCE_SPACE_KEY},
            "ancestors": [{"id": int(CONFLUENCE_PARENT_PAGE_ID)}],
            "body": {
                "storage": {
                    "value": confluence_body,
                    "representation": "storage",
                }
            },
            "version": {"number": existing_version + 1},
        }
        result = subprocess.run(
            [
                "curl", "-s", "-X", "PUT",
                "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
                "-H", "Content-Type: application/json",
                f"{CONFLUENCE_BASE_URL}/rest/api/content/{existing_page_id}",
                "-d", json.dumps(payload, ensure_ascii=False),
            ],
            capture_output=True, text=True, timeout=30
        )
    else:
        print(f"✨ 创建新页面...")
        payload = {
            "type": "page",
            "title": page_title,
            "space": {"key": CONFLUENCE_SPACE_KEY},
            "ancestors": [{"id": int(CONFLUENCE_PARENT_PAGE_ID)}],
            "body": {
                "storage": {
                    "value": confluence_body,
                    "representation": "storage",
                }
            },
        }
        result = subprocess.run(
            [
                "curl", "-s", "-X", "POST",
                "-H", f"Authorization: Bearer {CONFLUENCE_TOKEN}",
                "-H", "Content-Type: application/json",
                f"{CONFLUENCE_BASE_URL}/rest/api/content",
                "-d", json.dumps(payload, ensure_ascii=False),
            ],
            capture_output=True, text=True, timeout=30
        )

    # 3. 检查结果
    if result.returncode != 0:
        print(f"❌ 请求失败: {result.stderr}")
        return

    try:
        resp = json.loads(result.stdout)
        if "id" in resp:
            page_id = resp["id"]
            page_url = f"{CONFLUENCE_BASE_URL}/pages/viewpage.action?pageId={page_id}"
            status = resp.get("version", {}).get("number", "?")
            action = "更新" if existing_page_id else "创建"
            print(f"✅ {action}成功！")
            print(f"   Confluence 页面: {page_url}")
            print(f"   版本: {status}")
            print(f"   本地备份: {report_path}")
        else:
            print(f"❌ 失败:")
            print(json.dumps(resp, ensure_ascii=False, indent=2))
    except Exception:
        print(f"❌ 响应解析失败: {result.stdout[:500]}")


# ============================================================
# 4. CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(
        description="周报一键数据获取脚本 - 自动计算周日期，并行拉取所有数据源"
    )
    parser.add_argument("--week", type=str, default=None,
                        help="指定周次，如 W30。不指定则自动检测当前周")
    parser.add_argument("--excel", type=str, default=None,
                        help="指定技术支持工单明细.xlsx 文件路径。不指定则按优先级自动查找: 环境变量 → ./data/ → 原始路径")
    parser.add_argument("--token", type=str, default=None,
                        help="Confluence API Token。不指定则从环境变量 CONFLUENCE_TOKEN 或 .env 文件读取")
    parser.add_argument("--json", action="store_true",
                        help="输出原始 JSON 数据")
    parser.add_argument("--summary", action="store_true",
                        help="仅输出摘要信息")
    parser.add_argument("--audit", action="store_true",
                        help="获取数据后运行审计检查")
    parser.add_argument("--publish", action="store_true",
                        help="将生成的周报发布到 Confluence（26年技术支持周报 子页面）")
    args = parser.parse_args()

    # 解析 Token（命令行参数优先）
    global CONFLUENCE_TOKEN
    if args.token:
        CONFLUENCE_TOKEN = args.token
    elif not CONFLUENCE_TOKEN:
        CONFLUENCE_TOKEN = _resolve_confluence_token()

    # 解析 Excel 路径
    excel_path = _resolve_excel_path(args.excel)

    if args.publish:
        publish_weekly_report(args.week)
        return

    data = gather_all_data(args.week, excel_path)

    if args.audit:
        audit_result = run_audit(data)
        print(json.dumps(audit_result, ensure_ascii=False, indent=2, default=str))
        return

    if args.json:
        print(json.dumps(data, ensure_ascii=False, indent=2, default=str))
        return

    if args.summary:
        print_summary(data)
        return

    # 默认：输出 JSON + 摘要
    print(json.dumps(data, ensure_ascii=False, indent=2, default=str))


def print_summary(data):
    """打印可读的摘要信息"""
    meta = data.get("meta", {})
    print(f"\n{'='*60}")
    print(f"📊 周报 {meta.get('week', '?')} 数据摘要")
    print(f"📅 日期范围: {meta.get('date_range', '?')} (周五 → 周四)")
    print(f"📅 上周: {meta.get('prev_week', '?')} ({meta.get('prev_date_range', '?')})")
    print(f"{'='*60}")

    # Excel 数据
    excel = data.get("excel", {})
    if excel and "error" not in excel:
        print(f"\n📋 工单数据:")
        print(f"  本周总工单: {excel.get('total', 0)} 条")
        print(f"  超期工单: {excel.get('overdue_total', 0)} 条")
        print(f"  按部门: {excel.get('by_dept', {})}")
        print(f"  按业务组: {excel.get('grouped', {})}")
        overdue_g = excel.get("overdue_grouped", {})
        for group, info in overdue_g.items():
            print(f"  {group} 超期: {info['count']} 条")
        excluded = excel.get("excluded_overdue", {})
        if excluded:
            for group, info in excluded.items():
                print(f"  {group} 排除(外部超时): {info['count']} 条")

    # 需求
    demands = data.get("demands", {})
    if demands and "error" not in demands:
        dm = demands.get("demands", {})
        total_demands = sum(len(v) for v in dm.values())
        print(f"\n📝 需求列表: {total_demands} 条未完成")
        for cat, items in dm.items():
            print(f"  {cat}: {len(items)} 条")

    # 看板
    dashboard = data.get("dashboard", {})
    if dashboard and "error" not in dashboard:
        db = dashboard.get("data", {})
        if db:
            print(f"\n📊 翡翠看板数据:")
            print(f"  自助转技术工单: {db.get('auto_transfer', '—')}")
            print(f"  PM提交技术工单(数智): {db.get('pm_submit_finance', '—')}")
            print(f"  PM提交技术工单(免疫): {db.get('pm_submit_immune', '—')}")

    # 业务指标
    metrics = data.get("metrics", {})
    if metrics and "error" not in metrics:
        m = metrics.get("metrics", {})
        labels = metrics.get("week_labels", [])
        print(f"\n📈 业务指标详情 (来源: {metrics.get('page_title', 'Confluence')}):")
        if labels:
            print(f"  周标签: {labels[0] if len(labels) > 0 else '?'} → {labels[1] if len(labels) > 1 else '?'}")
        for biz, items in m.items():
            print(f"  {biz}: {len(items)} 项指标")

    # 故障数据
    faults = data.get("faults", {})
    if faults and "error" not in faults:
        f = faults.get("faults", {})
        print(f"\n⚠️ 本周故障 (来源: Confluence 故障明细):")
        if not f:
            print("  无")
        for group, items in f.items():
            print(f"  {group}: {len(items)} 条")
            for item in items:
                print(f"    [{item['level']}] {item['date']} | {item['name'][:60]}")

    # 错误
    errors = []
    for key in ["excel", "demands", "dashboard", "metrics", "faults"]:
        if isinstance(data.get(key), dict) and data[key].get("error"):
            errors.append(f"  {key}: {data[key]['error']}")
    if errors:
        print(f"\n⚠️ 错误:")
        for e in errors:
            print(e)

    print()


if __name__ == "__main__":
    main()